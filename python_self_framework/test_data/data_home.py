import openpyxl


class HomeData:
    test_home_data = [{"firstName" : "Jayden", "email" : "Seonggon@mail.com", "gender" : "Male"}, {"fitstName" : "Lala", "email" : "Rara@mail.com", "gender" : "Female"}]

    @staticmethod
    def get_test_data(test_case_name):
        book = openpyxl.load_workbook("PythonDemo.xlsx")
        sheet = book.active
        Dict = {}
        for i in range(1,
                       sheet.max_row + 1):  # to get rows [{"firstName" : "Jayden", "email" : "Seonggon@mail.com", "gender" : "Male"}, {"fitstName" : "Lala", "email" : "Rara@mail.com", "gender" : "Female"}]
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # to get columns
                    # Dict["lastname"]="shetty"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]